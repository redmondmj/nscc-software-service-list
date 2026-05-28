import json
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

JSON_PATH = "software_inventory.json"
EXCEL_PATH = "software_service_list.xlsx"
REPORT_PATH = "software_service_report.md"

def load_data():
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def generate_excel(data):
    print("Generating formatted Excel spreadsheet...")
    df = pd.DataFrame(data)
    
    # Write to Excel
    df.to_excel(EXCEL_PATH, index=False)
    
    # Load with openpyxl to apply styling
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    ws.title = "Software Registry"
    
    # Styles
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # NSCC Premium Navy
    white_bold_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Segoe UI", size=10)
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Format Headers
    ws.row_dimensions[1].height = 28
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = navy_fill
        cell.font = white_bold_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
    # Format Data Rows
    for row_idx in range(2, len(data) + 2):
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = regular_font
            cell.border = thin_border
            
            # Alignments based on content
            col_name = df.columns[col_idx - 1]
            if col_name in ["Version", "License Type (Perpetual, Subscription, Client/Server,Free, Open Source, etc.)", "Scope  (lab, faculty, student, cloud, all)", "Is this a software or service that is already being used?", "Cost (only if known)", "Number of Licenses/Users (Only if known)"]:
                cell.alignment = center_alignment
            else:
                cell.alignment = cell_alignment
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Check length of cell value
            if cell.value:
                # Add extra padding for headers or wrap text
                val_len = len(str(cell.value))
                if cell.row == 1:
                    val_len = min(val_len, 25) # Cap header size in width calculation to allow wrapping
                max_len = max(max_len, val_len)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    wb.save(EXCEL_PATH)
    print(f"Success! Formatted spreadsheet saved to {EXCEL_PATH}")

def generate_report(data):
    print("Generating comprehensive Markdown report...")
    
    # Categorize software
    categories = {
        "Virtualization & Hypervisors": [],
        "Cloud Computing Platforms": [],
        "Operating Systems (Server & Client)": [],
        "Network Services & Infrastructure": [],
        "DevOps & Automation Tools": [],
        "Security & Penetration Testing Tools": [],
        "Development & Programming Utilities": [],
        "Collaboration & Productivity Services": []
    }
    
    inferred_list = []
    future_list = []
    
    for item in data:
        name = item["Software Name"]
        notes = item["Notes/Comments/Special Requirements"]
        
        # Check if future or inferred
        if "(Proposed addition" in notes:
            future_list.append(item)
            continue
        elif "(Inferred program" in notes:
            inferred_list.append(item)
            continue
            
        # Categorize core items
        name_lower = name.lower()
        if "vmware" in name_lower or "hyper-v" in name_lower or "proxmox" in name_lower or "virtualbox" in name_lower or "viewer" in name_lower:
            categories["Virtualization & Hypervisors"].append(item)
        elif "aws" in name_lower or "azure" in name_lower or "cloud" in name_lower:
            categories["Cloud Computing Platforms"].append(item)
        elif "windows server" in name_lower or "windows 1" in name_lower or "linux" in name_lower or "adk" in name_lower or "winpe" in name_lower or "windows pe" in name_lower or "winre" in name_lower or "dism" in name_lower or "rufus" in name_lower:
            categories["Operating Systems (Server & Client)"].append(item)
        elif "dhcp" in name_lower or "dns" in name_lower or "iis" in name_lower or "apache" in name_lower or "nginx" in name_lower or "samba" in name_lower or "linkrunner" in name_lower or "tftp" in name_lower or "wifiman" in name_lower:
            categories["Network Services & Infrastructure"].append(item)
        elif "ansible" in name_lower or "terraform" in name_lower or "git" in name_lower or "kubernetes" in name_lower or "docker" in name_lower or "sccm" in name_lower or "mcm" in name_lower or "puppet" in name_lower or "chef" in name_lower or "ninite" in name_lower:
            categories["DevOps & Automation Tools"].append(item)
        elif "wireshark" in name_lower or "nmap" in name_lower or "metasploit" in name_lower or "nessus" in name_lower or "pfsense" in name_lower or "opensense" in name_lower or "defender" in name_lower:
            categories["Security & Penetration Testing Tools"].append(item)
        elif "vs code" in name_lower or "python" in name_lower or "mysql" in name_lower or "postgresql" in name_lower or "mongodb" in name_lower or "putty" in name_lower or "draw.io" in name_lower or "tera term" in name_lower or "sysinternals" in name_lower or "ollama" in name_lower or "windirstat" in name_lower or "admin center" in name_lower or "sql server" in name_lower or "visual studio" in name_lower or "sandbox" in name_lower:
            categories["Development & Programming Utilities"].append(item)
        else:
            categories["Collaboration & Productivity Services"].append(item)
            
    # Compile markdown text
    md = []
    md.append("# IT Systems Management & Security (ITSYSSEC) Program - Software & Service Report")
    md.append("\nThis report compiles all the software tools, cloud services, and hardware utilities used in the two-year **IT Systems Management and Security** diploma program at NSCC. It maps each software to its target courses and details licensing types, usage scope, and deployment settings.")
    
    md.append("\n## Core Software & Service Catalog")
    
    for cat_name, items in categories.items():
        if not items:
            continue
        md.append(f"\n### {cat_name}")
        md.append("| Software/Service Name | Courses Used In | License Type | Scope | Notes |")
        md.append("| --- | --- | --- | --- | --- |")
        for item in items:
            courses = item["What is the specific course/module/project you are using this software for?"]
            license_t = item["License Type (Perpetual, Subscription, Client/Server,Free, Open Source, etc.)"]
            scope = item["Scope  (lab, faculty, student, cloud, all)"]
            notes = item["Notes/Comments/Special Requirements"]
            md.append(f"| **{item['Software Name']}** | {courses} | {license_t} | *{scope}* | {notes} |")
            
    md.append("\n## Inferred Software & assumed Prerequisites")
    md.append("These tools are essential prerequisites and daily administrative utilities assumed for coursework but not explicitly registered in standard syllabus documents:")
    md.append("| Software Name | Vendor | Usage & Assumptions | Courses |")
    md.append("| --- | --- | --- | --- |")
    for item in inferred_list:
        courses = item["What is the specific course/module/project you are using this software for?"]
        notes = item["Notes/Comments/Special Requirements"].replace(" (Inferred program prerequisite)", "")
        md.append(f"| **{item['Software Name']}** | {item['Vendor/Provider']} | {notes} | {courses} |")
        
    md.append("\n## Future Software Requirements (Next 12 Months)")
    md.append("These technologies are proposed for integration into the curriculum over the next academic year to keep pace with industry developments in DevSecOps and cloud automation:")
    md.append("| Software Name | Vendor | Rationale for Curriculum Integration | Target Course |")
    md.append("| --- | --- | --- | --- |")
    for item in future_list:
        courses = item["What is the specific course/module/project you are using this software for?"]
        notes = item["Notes/Comments/Special Requirements"].replace(" (Proposed addition for upcoming academic year)", "")
        md.append(f"| **{item['Software Name']}** | {item['Vendor/Provider']} | {notes} | {courses} |")
        
    # Save Report
    with open(REPORT_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(md))
        
    print(f"Success! Report saved to {REPORT_PATH}")

def main():
    data = load_data()
    generate_excel(data)
    generate_report(data)
    print("All outputs generated successfully!")

if __name__ == "__main__":
    main()
